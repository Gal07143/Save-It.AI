"""Reports and Export API endpoints."""
from typing import List, Optional
from datetime import datetime
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from backend.app.core.database import get_db
from backend.app.models import Site, Meter, Bill, Asset

router = APIRouter(prefix="/api/v1", tags=["reports"])


@router.get("/export/sites")
def export_sites_excel(db: Session = Depends(get_db)):
    """Export sites to Excel file."""
    sites = db.query(Site).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sites"
    
    headers = ["ID", "Name", "Address", "City", "Country", "Timezone", "Created At"]
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row, site in enumerate(sites, 2):
        ws.cell(row=row, column=1, value=site.id)
        ws.cell(row=row, column=2, value=site.name)
        ws.cell(row=row, column=3, value=site.address or "")
        ws.cell(row=row, column=4, value=site.city or "")
        ws.cell(row=row, column=5, value=site.country or "")
        ws.cell(row=row, column=6, value=site.timezone)
        ws.cell(row=row, column=7, value=str(site.created_at))
    
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sites_export.xlsx"}
    )


@router.get("/export/meters")
def export_meters_excel(site_id: Optional[int] = None, db: Session = Depends(get_db)):
    """Export meters to Excel file."""
    query = db.query(Meter)
    if site_id:
        query = query.filter(Meter.site_id == site_id)
    meters = query.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Meters"
    
    headers = ["ID", "Meter ID", "Name", "Site ID", "Manufacturer", "Model", "Serial Number", "Active"]
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    for row, meter in enumerate(meters, 2):
        ws.cell(row=row, column=1, value=meter.id)
        ws.cell(row=row, column=2, value=meter.meter_id)
        ws.cell(row=row, column=3, value=meter.name)
        ws.cell(row=row, column=4, value=meter.site_id)
        ws.cell(row=row, column=5, value=meter.manufacturer or "")
        ws.cell(row=row, column=6, value=meter.model or "")
        ws.cell(row=row, column=7, value=meter.serial_number or "")
        ws.cell(row=row, column=8, value="Yes" if meter.is_active else "No")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=meters_export.xlsx"}
    )


@router.get("/export/bills")
def export_bills_excel(site_id: Optional[int] = None, db: Session = Depends(get_db)):
    """Export bills to Excel file."""
    query = db.query(Bill)
    if site_id:
        query = query.filter(Bill.site_id == site_id)
    bills = query.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bills"
    
    headers = ["ID", "Site ID", "Provider", "Bill Period Start", "Bill Period End", 
               "Total kWh", "Peak kW", "Total Amount", "Currency", "Validated"]
    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    for row, bill in enumerate(bills, 2):
        ws.cell(row=row, column=1, value=bill.id)
        ws.cell(row=row, column=2, value=bill.site_id)
        ws.cell(row=row, column=3, value=bill.provider_name or "")
        ws.cell(row=row, column=4, value=str(bill.period_start))
        ws.cell(row=row, column=5, value=str(bill.period_end))
        ws.cell(row=row, column=6, value=bill.total_kwh or 0)
        ws.cell(row=row, column=7, value=bill.demand_kw or 0)
        ws.cell(row=row, column=8, value=bill.total_amount or 0)
        ws.cell(row=row, column=9, value=bill.currency)
        ws.cell(row=row, column=10, value="Yes" if bill.is_validated else "No")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bills_export.xlsx"}
    )


@router.get("/reports/site-summary/{site_id}")
def generate_site_report_pdf(site_id: int, db: Session = Depends(get_db)):
    """Generate PDF report for a site."""
    site = db.query(Site).filter(Site.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    
    meters = db.query(Meter).filter(Meter.site_id == site_id).all()
    bills = db.query(Bill).filter(Bill.site_id == site_id).order_by(Bill.period_end.desc()).limit(12).all()
    assets = db.query(Asset).filter(Asset.site_id == site_id).all()
    
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=24, textColor=colors.HexColor('#1E40AF'))
    story.append(Paragraph(f"SAVE-IT.AI - Site Report", title_style))
    story.append(Spacer(1, 20))
    story.append(Paragraph(f"<b>Site:</b> {site.name}", styles['Normal']))
    story.append(Paragraph(f"<b>Address:</b> {site.address or 'N/A'}, {site.city or ''}, {site.country or ''}", styles['Normal']))
    story.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 30))
    
    story.append(Paragraph("Site Summary", styles['Heading2']))
    summary_data = [
        ["Metric", "Value"],
        ["Total Assets", str(len(assets))],
        ["Active Meters", str(len([m for m in meters if m.is_active]))],
        ["Total Bills", str(len(bills))],
        ["Total Capacity (kW)", str(sum(a.rated_capacity_kw or 0 for a in assets))],
    ]
    summary_table = Table(summary_data, colWidths=[200, 200])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 30))
    
    if bills:
        story.append(Paragraph("Recent Bills", styles['Heading2']))
        bill_data = [["Period", "kWh", "Peak kW", "Amount", "Validated"]]
        for bill in bills[:6]:
            bill_data.append([
                f"{bill.period_start} - {bill.period_end}",
                f"{bill.total_kwh:,.0f}" if bill.total_kwh else "N/A",
                f"{bill.demand_kw:,.1f}" if bill.demand_kw else "N/A",
                f"{bill.currency} {bill.total_amount:,.2f}" if bill.total_amount else "N/A",
                "Yes" if bill.is_validated else "No"
            ])
        bill_table = Table(bill_data, colWidths=[120, 80, 80, 100, 70])
        bill_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        story.append(bill_table)
    
    doc.build(story)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=site_{site_id}_report.pdf"}
    )


@router.get("/reports/energy-analysis/{site_id}")
def generate_energy_analysis_pdf(site_id: int, db: Session = Depends(get_db)):
    """Generate energy analysis PDF report."""
    site = db.query(Site).filter(Site.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Site not found")
    
    bills = db.query(Bill).filter(Bill.site_id == site_id).order_by(Bill.period_end.desc()).limit(12).all()
    
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    story.append(Paragraph("Energy Analysis Report", styles['Title']))
    story.append(Spacer(1, 20))
    story.append(Paragraph(f"<b>Site:</b> {site.name}", styles['Normal']))
    story.append(Paragraph(f"<b>Report Period:</b> Last 12 Months", styles['Normal']))
    story.append(Spacer(1, 30))
    
    if bills:
        total_kwh = sum(b.total_kwh or 0 for b in bills)
        total_cost = sum(b.total_amount or 0 for b in bills)
        avg_monthly_kwh = total_kwh / len(bills)
        avg_monthly_cost = total_cost / len(bills)
        peak_demand = max(b.demand_kw or 0 for b in bills)
        
        story.append(Paragraph("Key Metrics", styles['Heading2']))
        metrics_data = [
            ["Metric", "Value"],
            ["Total Energy Consumption", f"{total_kwh:,.0f} kWh"],
            ["Total Energy Cost", f"${total_cost:,.2f}"],
            ["Average Monthly Consumption", f"{avg_monthly_kwh:,.0f} kWh"],
            ["Average Monthly Cost", f"${avg_monthly_cost:,.2f}"],
            ["Peak Demand", f"{peak_demand:,.1f} kW"],
            ["Average Cost per kWh", f"${total_cost/total_kwh:.4f}" if total_kwh > 0 else "N/A"],
        ]
        metrics_table = Table(metrics_data, colWidths=[200, 200])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        story.append(metrics_table)
    
    doc.build(story)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=energy_analysis_{site_id}.pdf"}
    )
